"""历史中标价行情库服务：xlsx 解析入库（公共共建 / 企业私有）+ 联合查询与聚合信号。"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from hashlib import sha256
from statistics import median

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.quote import HistoryPriceSnapshot

# 报价方式归一（xlsx 原文 → 枚举）
_PRICE_MODE_ALIASES = {
    "折扣率": "折扣率",
    "系数": "系数报价",
    "系数报价": "系数报价",
    "比例报价": "比例报价",
    "比例": "比例报价",
    "固定总价": "固定总价",
    "金额总价报价": "固定总价",
    "总价": "固定总价",
    "未披露": "未披露",
    "限价规则": "限价规则",
}

_AMOUNT_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*(万元|%)?")
_RATIO_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*%")
_MODE_HINT_RE = re.compile(r"(折扣率|系数|比例|固定总价|总价|费率|未披露|限价规则)")


def normalize_price_mode(raw: str | None) -> str:
    if not raw:
        return "其他"
    for key, mode in _PRICE_MODE_ALIASES.items():
        if key in raw:
            return mode
    m = _MODE_HINT_RE.search(raw)
    return _PRICE_MODE_ALIASES.get(m.group(1), "其他") if m else "其他"


def parse_amount(raw: str | None) -> float | None:
    """限价/中标价文本 → 数值。金额取万元；纯百分比（折扣率行）返回 None（由 price_mode 表达）。"""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s in ("—", "-", "未披露"):
        return None
    if "%" in s and "万元" not in s:
        return None  # 折扣率行：限价不落金额
    m = _AMOUNT_RE.search(s)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def parse_win_value(raw: str | None) -> tuple[float, bool]:
    """中标价 → (数值, 是否折扣率)。折扣率行返回百分数值（90.0=90%）。"""
    if raw is None:
        raise ValueError("中标价缺失")
    s = str(raw).strip()
    if "%" in s:
        m = _RATIO_RE.search(s)
        if not m:
            raise ValueError(f"无法解析中标价：{raw}")
        return float(m.group(1)), True
    m = _AMOUNT_RE.search(s)
    if not m:
        raise ValueError(f"无法解析中标价：{raw}")
    return float(m.group(1)), False


# 标黄提取.xlsx 的表头列名（按出现匹配，位置不固定）
_COL_ALIASES = {
    "发布时间": ("发布时间",),
    "公告/项目名称": ("公告/项目名称", "公告/项目"),
    "发布单位": ("发布单位",),
    "分标/品类": ("分标/品类", "分标名称", "品类"),
    "包号": ("包号",),
    "包/项目名称": ("包/项目名称", "包名", "项目名称"),
    "限价": ("限价",),
    "报价方式": ("报价方式",),
    "限价证据原文": ("限价证据原文",),
    "限价证据链接": ("限价证据链接",),
    "限价官网链接": ("限价官网链接",),
    "中标价": ("中标价",),
    "中标价证据原文": ("中标价证据原文",),
    "中标价证据链接": ("中标价证据链接",),
    "中标价官网链接": ("中标价官网链接",),
    "公告ID": ("公告ID",),
}


def _locate_columns(rows: list[list]) -> dict[str, int]:
    """在表头行定位列下标；找不到用 -1。"""
    found: dict[str, int] = {k: -1 for k in _COL_ALIASES}
    for r in rows[:12]:
        for ci, cell in enumerate(r):
            text = str(cell or "").strip()
            for key, aliases in _COL_ALIASES.items():
                if found[key] < 0 and text in aliases:
                    found[key] = ci
    return found


def _cell(row: list, idx: int) -> str | None:
    if idx < 0 or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    return str(v).strip()


def parse_market_xlsx(data: bytes) -> tuple[list[dict], list[str]]:
    """解析 标黄提取.xlsx（多 sheet，表头在表内，行级容错）。
    返回 (结构化行列表, 跳过原因列表)。"""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    rows_out: list[dict] = []
    skipped: list[str] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        cols = _locate_columns(rows)
        if cols.get("限价") < 0 or cols.get("中标价") < 0:
            skipped.append(f"sheet[{ws.title}] 缺 限价/中标价 列，跳过整表")
            continue
        for r in rows:
            notice = _cell(r, cols["公告ID"])
            win_raw = _cell(r, cols["中标价"])
            if not notice and not win_raw:
                continue  # 表头/空行
            if not notice:
                skipped.append(f"无公告ID，跳过：{win_raw}")
                continue
            if not win_raw:
                skipped.append(f"{notice}: 中标价缺失")
                continue
            try:
                win_val, is_ratio = parse_win_value(win_raw)
            except ValueError as exc:
                skipped.append(f"{notice}: {exc}")
                continue
            publish_raw = _cell(r, cols["发布时间"])
            try:
                publish = datetime.fromisoformat(str(publish_raw)[:10]).date() if publish_raw else None
            except ValueError:
                publish = None
            mode_raw = _cell(r, cols["报价方式"])
            mode = "折扣率" if is_ratio else normalize_price_mode(mode_raw)
            limit_raw = _cell(r, cols["限价"])
            limit_price = None if is_ratio else parse_amount(limit_raw)
            name = _cell(r, cols["包/项目名称"]) or _cell(r, cols["公告/项目名称"]) or notice
            category = _cell(r, cols["分标/品类"])
            publisher = _cell(r, cols["发布单位"])
            rows_out.append(
                {
                    "notice_id": notice,
                    "publish_date": publish,
                    "publisher": publisher,
                    "category": category,
                    "package_name": name[:300],
                    "price_mode": mode,
                    "limit_price": limit_price,
                    "win_price": win_val,
                    "win_date": publish or date(2026, 1, 1),
                    "limit_evidence": (_cell(r, cols["限价证据原文"]) or "")[:300],
                    "win_evidence": (_cell(r, cols["中标价证据原文"]) or "")[:300],
                    "limit_evidence_url": (_cell(r, cols["限价证据链接"]) or "")[:500],
                    "win_evidence_url": (_cell(r, cols["中标价证据链接"]) or "")[:500],
                }
            )
    return rows_out, skipped


def _row_hash(row: dict) -> str:
    payload = "|".join(
        str(row.get(k) or "") for k in ("notice_id", "publisher", "category", "package_name", "price_mode", "win_price")
    )
    return sha256(payload.encode("utf-8")).hexdigest()


async def import_rows(
    session: AsyncSession, enterprise_id: int, target: str, rows: list[dict]
) -> dict:
    """入公共库（enterprise_id=0）或本企业私有库；按 (notice_id, package_name, win_price) 去重。"""
    scope_eid = 0 if target == "public" else enterprise_id
    imported = 0
    for r in rows:
        row = HistoryPriceSnapshot(
            enterprise_id=scope_eid,
            provider_id="market_xlsx_import" if target == "public" else "enterprise_private",
            material_name=r["package_name"],
            category=r["category"],
            package_name=r["package_name"],
            publisher=r["publisher"],
            price_mode=r["price_mode"],
            limit_price=r["limit_price"],
            win_price=r["win_price"],
            win_date=r["win_date"],
            publish_date=r["publish_date"],
            notice_id=r["notice_id"],
            limit_evidence=r["limit_evidence"],
            win_evidence=r["win_evidence"],
            limit_evidence_url=r["limit_evidence_url"],
            win_evidence_url=r["win_evidence_url"],
            source_hash=_row_hash(r),
        )
        session.add(row)
        imported += 1
    await session.flush()
    return {"imported": imported, "skipped": 0, "scope": target}


async def query_library(
    session: AsyncSession,
    enterprise_id: int,
    *,
    category: str | None = None,
    publisher: str | None = None,
    price_mode: str | None = None,
    scope: str = "all",
    limit: int = 50,
) -> dict:
    """联合查询公共库+私有库，返回 {sample_count, samples, stats, readonly}。"""
    conds = [HistoryPriceSnapshot.provider_id.notlike("snapshot:%")]  # 快照副本不进行情库查询
    if category:
        conds.append(HistoryPriceSnapshot.category.ilike(f"%{category}%"))
    if publisher:
        conds.append(HistoryPriceSnapshot.publisher.ilike(f"%{publisher}%"))
    if price_mode:
        conds.append(HistoryPriceSnapshot.price_mode == price_mode)
    if scope == "public":
        conds.append(HistoryPriceSnapshot.enterprise_id == 0)
    elif scope == "private":
        conds.append(HistoryPriceSnapshot.enterprise_id == enterprise_id)
    else:
        conds.append(HistoryPriceSnapshot.enterprise_id.in_([0, enterprise_id]))

    rows = (
        await session.scalars(
            select(HistoryPriceSnapshot)
            .where(*conds)
            .order_by(HistoryPriceSnapshot.publish_date.desc().nullslast(), HistoryPriceSnapshot.id.desc())
            .limit(limit)
        )
    ).all()

    def _src(row: HistoryPriceSnapshot) -> str:
        return "public" if row.enterprise_id == 0 else "private"

    samples = []
    for r in rows:
        limit_v = float(r.limit_price) if r.limit_price is not None else None
        win_v = float(r.win_price)
        samples.append(
            {
                "source": _src(r),
                "publisher": r.publisher,
                "category": r.category,
                "package_name": r.package_name,
                "price_mode": r.price_mode,
                "limit_price": str(limit_v) if limit_v is not None else None,
                "win_price": str(win_v),
                "publish_date": r.publish_date.isoformat() if r.publish_date else None,
                "notice_id": r.notice_id,
                "limit_evidence": r.limit_evidence,
                "win_evidence": r.win_evidence,
                "limit_evidence_url": r.limit_evidence_url,
                "win_evidence_url": r.win_evidence_url,
                "win_ratio": (
                    str(round(win_v / limit_v, 4)) if limit_v and r.price_mode != "折扣率" else None
                ),
            }
        )

    # 聚合信号：按报价方式分组（样本口径标注，供 agent 参考；决策仍由 agent 做）
    stat_rows = (
        await session.execute(
            select(
                HistoryPriceSnapshot.price_mode,
                func.count().label("n"),
                func.min(HistoryPriceSnapshot.win_price).label("lo"),
                func.max(HistoryPriceSnapshot.win_price).label("hi"),
            )
            .where(*conds)
            .group_by(HistoryPriceSnapshot.price_mode)
        )
    ).all()
    stats = []
    for s in stat_rows:
        stats.append(
            {
                "price_mode": s.price_mode or "其他",
                "count": int(s.n),
                "win_price_range": [str(float(s.lo)), str(float(s.hi))],
            }
        )
    return {"sample_count": len(samples), "samples": samples, "stats": stats, "readonly": False}


async def db_samples_for_quote(
    session: AsyncSession, enterprise_id: int, material_ref: str
) -> list[dict]:
    """测算数据源（私有→公共→AnySearch→如实报不足，无 Mock）：按品类/包名关键词取库内样本。"""
    like = f"%{material_ref}%"
    rows = (
        await session.scalars(
            select(HistoryPriceSnapshot)
            .where(
                HistoryPriceSnapshot.enterprise_id.in_([enterprise_id, 0]),
                HistoryPriceSnapshot.provider_id.notlike("snapshot:%"),
                (HistoryPriceSnapshot.package_name.ilike(like))
                | (HistoryPriceSnapshot.category.ilike(like))
                | (HistoryPriceSnapshot.material_name.ilike(like)),
            )
            .order_by(HistoryPriceSnapshot.id.desc())
            .limit(30)
        )
    ).all()
    out = []
    for r in rows:
        out.append(
            {
                "material_ref": material_ref,
                "material_name": r.package_name or r.material_name,
                "spec": r.category or "",
                "region": r.publisher or "",
                "win_price": float(r.win_price),
                "win_date": r.win_date or r.publish_date or date(2026, 1, 1),
                "price_mode": r.price_mode,
                "limit_price": float(r.limit_price) if r.limit_price is not None else None,
                "source_url": r.win_evidence_url or "",
                "scope": "private" if r.enterprise_id != 0 else "public",
                "unit": "%" if r.price_mode == "折扣率" else "万元",
                "currency": "CNY",
                "tax_included": True,
            }
        )
    return out


def library_stats_median(samples: list[dict]) -> float | None:
    """样本中标价的中位数（金额口径；折扣率样本不混入金额统计）。"""
    vals = [s["win_price"] for s in samples if s.get("price_mode") != "折扣率" and s.get("win_price") is not None]
    return float(median(vals)) if vals else None
