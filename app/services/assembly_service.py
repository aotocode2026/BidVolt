"""成文工具链服务（新方案）：主会话自主成文的机制工具。

主会话经 MCP 调用本模块的机制原语，自主完成
"选底稿 → 列清单 → 切片 → 填空 → 追加 → 校验 → 封存 → 打包"：
- 机制保真：切片=底稿条目区间字节级复制；填空/追加=直接干净写入（无修订、无批注，输出即终稿）；校验=模板原文+替换链复算保真；
- 决策在主会话：填什么值、加什么内容、按什么顺序、何时封存打包，全部由主会话决定。

旧方案路径（build_response_package 整段服务端成文）不受影响。
"""

from __future__ import annotations

import logging
import time
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger(__name__)

# 内存切片仓：一个切片持有 (Document + 持久 _FillSession/editor)，保证替换记录全局连续。
# 容量上限 + TTL 防止无界增长（app 进程单 worker；MCP 会话串行调用）。
_SLICES: dict[str, dict[str, Any]] = {}
_SLICE_CAP = 40
_SLICE_TTL = 3600.0


def _prune_slices() -> None:
    now = time.time()
    for sid in [s for s, v in _SLICES.items() if now - v["created"] > _SLICE_TTL]:
        _SLICES.pop(sid, None)
    while len(_SLICES) > _SLICE_CAP:
        oldest = min(_SLICES, key=lambda s: _SLICES[s]["created"])
        _SLICES.pop(oldest, None)


def _slice(slice_id: str, task_id: int) -> dict[str, Any]:
    s = _SLICES.get(slice_id)
    if s is None:
        raise ValueError(f"切片不存在或已过期：{slice_id}（请重新 slice_template_item）")
    if int(s["task_id"]) != int(task_id):
        raise ValueError("切片不属于当前任务")
    return s


async def list_draft_candidates(session: AsyncSession, enterprise_id: int, project_id: int) -> dict:
    """候选底稿：按内容分级排序（含"响应文件格式"章优先），返回列表+推荐 file_id。
    主会话可据此选择底稿（也允许指定其他 docx）。"""
    from sqlalchemy import select as sa_select

    from app.models.doc import DocBlock
    from app.models.file import FileObject

    rows = (
        await session.scalars(
            sa_select(FileObject).where(
                FileObject.project_id == int(project_id),
                FileObject.enterprise_id == enterprise_id,
                FileObject.is_deleted.is_(False),
                FileObject.owner_type == 2,
            )
        )
    ).all()
    docx_rows = [f for f in rows if (f.ext or "").strip(".").lower() == "docx"]
    texts: dict[int, str] = {}
    for f in docx_rows:
        bs = (
            await session.scalars(
                sa_select(DocBlock.text_content).where(DocBlock.file_id == f.id)
            )
        ).all()
        texts[f.id] = "\n".join(b or "" for b in bs)[:200000]

    def _rank(f) -> tuple[int, str]:
        t = texts.get(f.id) or ""
        if "响应文件格式" in t:
            return (3, "含《响应文件格式》章（完整采购文件）")
        if "商务文件" in t and "技术文件" in t:
            return (2, "含商务/技术文件章节")
        if "商务文件" in t or "技术文件" in t:
            return (1, "含商务或技术文件章节")
        return (0, "其他 docx")

    candidates = []
    for f in docx_rows:
        rank, reason = _rank(f)
        candidates.append(
            {
                "file_id": f.id,
                "name": f.original_name or f.storage_name or "",
                "size_bytes": f.size_bytes or 0,
                "rank": rank,
                "reason": reason,
            }
        )
    candidates.sort(key=lambda c: (c["rank"], c["size_bytes"]), reverse=True)
    return {"candidates": candidates, "recommended_file_id": candidates[0]["file_id"] if candidates else None}


async def get_template_outline(session: AsyncSession, enterprise_id: int, project_id: int) -> dict:
    """模板清单：《响应文件格式》条目（doc_template），按 价格/商务/技术 分组排序。
    每项带 req_id（后续 slice_template_item 的 item_ref）与 is_file_item 标记：
    is_file_item=true = 真实成文条目（（一）（二）…编号条目，必须逐份成文），
    is_file_item=false = 结构/指引行（分册标题、上传路径说明等，不单独成文）。"""
    import re as _re

    from sqlalchemy import select as sa_select

    from app.models.requirement import Requirement

    rows = (
        await session.scalars(
            sa_select(Requirement).where(
                Requirement.enterprise_id == enterprise_id,
                Requirement.project_id == project_id,
                Requirement.current.is_(True),
                Requirement.req_type == "doc_template",
            )
        )
    ).all()
    items_by_role: dict[str, list[dict]] = {"price": [], "business": [], "technical": []}
    for r in rows:
        role = (r.structured or {}).get("role")
        if role in items_by_role:
            order = (r.structured or {}).get("order", 0)
            try:
                order = int(order)
            except (TypeError, ValueError):
                order = 99999
            title = (r.content or "").split("\n")[0].strip()
            # 真实成文条目 = 全角中文数字编号（（一）（二）…）的短标题行
            # （与底稿 outlineLvl=2 条目一致，阈值同切分器的短标题标准）；
            # 正文条款（（1）…、以冒号结尾的长句）与"标题+上传路径说明"登记行都不算。
            is_item = bool(
                _re.match(r"^[（(]\s*[一二三四五六七八九十百]+\s*[）)]", title)
                and len(title) <= 24
                and not title.endswith(("：", ":"))
            )
            items_by_role[role].append(
                {
                    "req_id": r.id,
                    "title": title[:80],
                    "order": order,
                    "is_file_item": is_item,
                }
            )
    for role in items_by_role:
        items_by_role[role].sort(key=lambda x: (isinstance(x["order"], int), x["order"]))
    return {"items_by_role": items_by_role, "counts": {k: len(v) for k, v in items_by_role.items()}}


async def create_slice(
    session: AsyncSession,
    enterprise_id: int,
    project_id: int,
    task_id: int,
    file_id: int,
    req_id: int,
) -> dict:
    """切片：把底稿中该条目区间原文整段复制为独立文档骨架，存入内存切片仓。
    返回 slice_id 与定位状态（未定位时按解析清单内容重建并如实标注）。"""
    import secrets

    from sqlalchemy import select as sa_select

    from app.models.file import FileObject
    from app.models.requirement import Requirement
    from app.services import export_service
    from app.services.storage import StorageProvider

    row = await session.scalar(
        sa_select(Requirement).where(
            Requirement.id == int(req_id),
            Requirement.enterprise_id == enterprise_id,
            Requirement.project_id == int(project_id),
            Requirement.current.is_(True),
            Requirement.req_type == "doc_template",
        )
    )
    if row is None:
        raise ValueError(f"模板条目不存在：req_id={req_id}")
    fobj = await session.get(FileObject, int(file_id))
    if fobj is None or fobj.enterprise_id != enterprise_id or fobj.is_deleted or (fobj.ext or "").strip(".").lower() != "docx":
        raise ValueError(f"底稿文件不可用：file_id={file_id}")
    source_path = StorageProvider().open(fobj.bucket, fobj.object_key)

    matched_title, elements = export_service.locate_item_with_heading(source_path, row)
    doc, located = export_service._build_item_document(source_path, row, elements)

    # 底稿原文（校验基准）：与 check_doc_fidelity 同侧同算法（canonical_text），
    # 剔除绘图内部坐标数字等非正文噪声，避免锚定图（印模）导致合法切片误报不忠实
    import zipfile as _zip

    from lxml import etree as _etree

    with _zip.ZipFile(source_path) as zf:
        src_xml = _etree.fromstring(zf.read("word/document.xml"))
    source_text = export_service.canonical_text(src_xml)
    # 切片模板原文（忠实性校验基准之一：干净成文后无法靠修订层区分原文与新增，
    # 改由「原文+替换链」复算保真，此值即替换链起点）
    original_text = export_service.canonical_text(doc.element)

    _prune_slices()
    slice_id = "s" + secrets.token_hex(6)
    req_title = (row.content or "").split("\n")[0].strip()
    _SLICES[slice_id] = {
        "doc": doc,
        "sess": None,
        "source_text": source_text,
        "original_text": original_text,
        "file_id": int(file_id),
        "req_id": int(req_id),
        "title": req_title,
        "matched_title": (matched_title or "").strip(),
        "verified": False,  # 信息信号：是否通过过 verify（seal 回执带 was_verified 供 agent 自查）
        "task_id": int(task_id),
        "project_id": int(project_id),
        "enterprise_id": int(enterprise_id),
        "created": time.time(),
    }
    # 身份信号：matched_title 是底稿里实际匹配到的条目标题，
    # 主会话必须比对它与请求条目标题一致（不一致=串区，立即重切）
    warn = (
        None
        if located
        else "该条目在底稿中未定位到原文区间，按解析清单内容重建（可能不完整），请对照采购文件原件核对。"
    )
    return {
        "slice_id": slice_id,
        "located": located,
        "matched_title": (matched_title or "")[:80],
        "req_title": req_title[:80],
        "file_id": int(file_id),
        "req_id": int(req_id),
        # 表格清册：让 agent 看到本切片有哪些模板表格（坐标/行数/列数/表头），
        # 才能用 table_fills 把内容填进模板自身的表格
        "tables": export_service.tables_inventory(doc.element),
        # 切片内容预览：agent 填之前先看到这段原文（信息信号，机制只描述）
        "slice_preview": {
            "text_head": "".join(
                t.text or "" for t in doc.element.iter(f"{{{export_service._W_NS}}}t")
            )[:600],
        },
        "warn": warn,
    }


def _ensure_sess(s, fields: dict) -> Any:
    from app.services import export_service

    if s["sess"] is None:
        s["sess"] = export_service._FillSession(
            s["doc"],
            str(fields.get("buyer") or "").strip(),
            str(fields.get("project_name") or "").strip(),
            str(fields.get("supplier") or "").strip(),
            str(fields.get("tender_no") or "").strip(),
            label_values=(fields.get("label_values") if isinstance(fields.get("label_values"), dict) else {}),
            values=(fields.get("values") if isinstance(fields.get("values"), dict) else {}),
        )
    else:
        sess = s["sess"]
        for key, attr in (
            ("buyer", "buyer"),
            ("project_name", "project_name"),
            ("supplier", "supplier"),
            ("tender_no", "tender_no"),
        ):
            if fields.get(key) is not None:
                setattr(sess, attr, str(fields[key]).strip())
        if isinstance(fields.get("label_values"), dict) and fields["label_values"]:
            sess.set_label_values(fields["label_values"])
        if isinstance(fields.get("values"), dict) and fields["values"]:
            sess.set_label_values(fields["values"])
    return s["sess"]


def fill_slice(slice_id: str, task_id: int, fields: dict | None, fills: list[dict] | None, table_fills: list[dict] | None = None) -> dict:
    """填空：标准字段（buyer/project_name/supplier/tender_no）→ 带标签空位规则（无资料原位【待补充】），
    再按主会话给出的显式 fills=[{find,value,comment}] 定向替换；
    table_fills=[{table,row,col,value,comment}] 按 agent 决定的坐标填单元格（工具只给表格清册，
    填哪行哪列由 agent 决定）。全部直接干净写入正文（无修订、无批注）。"""
    s = _slice(slice_id, task_id)
    sess = _ensure_sess(s, fields or {})
    sess.apply_to_doc()
    from app.services import export_service  # noqa: PLC0415

    n_fills = 0
    fills_results: list[dict] = []
    for f in fills or []:
        find = str(f.get("find") or "")
        if not find:
            continue
        value = str(f.get("value") or "")
        comment = str(f.get("comment") or "") or None
        n = export_service.replace_text_tracked(sess.editor, find, value, comment)
        n_fills += n
        res: dict = {
            "find": find[:60],
            "value": value[:60],
            "replaced": n,
            "found": n > 0,
        }
        if n == 0:
            # 信号（非裁决）：最常见原因是 find 串跨了段落——替换只按段落匹配。
            res["hint"] = "find 只按段落匹配：若 find 串跨两个段落会匹配不到，请拆成单段 find 重试"
        fills_results.append(res)
    n_table = 0
    table_results: list[dict] = []

    def _idx(v, default: int = -1) -> int:
        try:
            return int(v)
        except (TypeError, ValueError):
            return default

    for tf in table_fills or []:
        # 注意：0 是合法坐标（表 0/行 0/列 0），不能用 `or -1` 兜底（曾把 0 腐蚀成 -1，
        # 导致所有以 0 开头的坐标全部越界，agent 误判"机制未生效"而放弃填原表）
        ti = _idx(tf.get("table"))
        ri = _idx(tf.get("row"))
        ci = _idx(tf.get("col"))
        err = ""
        try:
            ok = sess.fill_table_cell(ti, ri, ci, str(tf.get("value") or ""), str(tf.get("comment") or "") or None)
        except Exception as exc:  # noqa: BLE001
            ok = False
            err = f"{type(exc).__name__}: {exc}"
        if ok:
            n_table += 1
        else:
            n_tables = len(sess.doc.tables)
            err = err or f"坐标越界：本切片共 {n_tables} 张表（table 索引 0-{n_tables - 1}），"
            if n_tables and 0 <= ti < n_tables:
                rows, cols = len(sess.doc.tables[ti].rows), len(sess.doc.tables[ti].rows[0].cells)
                err += f"表 {ti} 为 {rows} 行 × {cols} 列"
        table_results.append({"table": ti, "row": ri, "col": ci, "ok": ok, "error": err})
    s["verified"] = False  # 信息信号：内容有改动，was_verified 置否（agent 应重验后再封存）
    remaining = sess.remaining_blanks()
    return {
        "slice_id": slice_id,
        "explicit_fills": n_fills,
        "fills_results": fills_results,
        "table_fills_done": n_table,
        "table_fills_results": table_results,
        "fields_used": {k: str(v) for k, v in (fields or {}).items() if v},
        # 填完即反馈：本文档还有哪些空位没填（逐项标签+上下文），主会话逐项清零到只剩客户独占数据
        "remaining_blanks": remaining,
        "remaining_count": len(remaining),
        # 表格还有哪些全空数据行（描述信号，位置决定权在 agent）
        "empty_table_rows": export_service.empty_table_rows(sess.doc.element),
    }


def append_slice(slice_id: str, task_id: int, nodes: list[dict] | None, comment: str | None, heading: str | None = None) -> dict:
    """追加撰写内容：直接追加为正文（无修订无批注；节点形状兼容段落/标题/表格/裸字符串）。
    heading 由主会话按投标文体自定（方案类条目的正文追加用）；不传时用中性默认"响应内容"。"""
    s = _slice(slice_id, task_id)
    sess = _ensure_sess(s, {})
    sess.append_supplement(
        nodes or [],
        heading_text=(str(heading).strip() if heading else "响应内容"),
    )
    s["verified"] = False  # 信息信号：内容有改动，was_verified 置否（agent 应重验后再封存）
    return {"slice_id": slice_id, "appended_nodes": len(nodes or []), "heading": heading or "响应内容"}


def verify_slice(slice_id: str, task_id: int) -> dict:
    """忠实性校验（干净成文口径）：模板段落「原文+替换链」复算保真（不被直接改写/删除）
    + 切片模板原文逐字⊂底稿。不过时返回差异片段。
    附带身份信息（req_title=请求条目、matched_title=实际绑定条目）供主会话比对；
    通过后置 verified 标记（seal 回执的 was_verified 信号，供验收子 agent 核对）。"""
    s = _slice(slice_id, task_id)
    from app.services import export_service

    sess = s.get("sess")
    r = export_service.check_doc_fidelity(
        s["doc"],
        s["source_text"],
        editor=(sess.editor if sess is not None else None),
        original_text=s.get("original_text"),
    )
    r["slice_id"] = slice_id
    r["req_title"] = s.get("title") or ""
    r["matched_title"] = s.get("matched_title") or ""
    # 统一社会信用代码格式校验（R8 教训：营业执照 OCR 曾漏一位成 17 位，进包=硬伤）：
    # 全文扫 91 开头的独立数字字母串，非 18 位一律判不过，由 agent 回修。
    import re as _re_code

    _doc_text_c = export_service._elem_text(s["doc"].element)
    _code_issues = [
        _m.group(0)
        for _m in _re_code.finditer(r"(?<![0-9A-Z])91[0-9A-Z]{15,17}(?![0-9A-Z])", _doc_text_c)
        if len(_m.group(0)) != 18
    ]
    if _code_issues:
        r["ok"] = False
        r["credit_code_issues"] = _code_issues
    r["passed"] = bool(r["ok"])
    s["verified"] = bool(r["ok"])
    if not r["ok"]:
        s["verified"] = False
    # 校验时同步反馈：还有哪些空位没填（信息信号）
    r["remaining_blanks"] = s["sess"].remaining_blanks() if s.get("sess") is not None else export_service._scan_remaining(s["doc"].element)
    r["remaining_count"] = len(r["remaining_blanks"])
    return r


async def seal_slice(
    session: AsyncSession,
    slice_id: str,
    task_id: int,
    dir_name: str,
    filename: str,
) -> dict:
    """封存：生成条目 docx 并落产物表。
    服务端不做硬性拦截（产品决定：合规性由主会话+验收/评审子 agent 保证），
    仅返回信息信号（matched_title/verified）供主会话自查。"""
    from app.services.export_service import _clean_item_name, _safe_filename

    s = _slice(slice_id, task_id)
    sess = _ensure_sess(s, {})
    # 未显式填空也走一遍规则：带标签空位无资料原位【待补充】（诚实标注）
    sess.apply_to_doc()
    data = sess.finish()

    from app.models.agent import AgentArtifact
    from app.services.task_service import _set_rls_context  # noqa: PLC0415

    await _set_rls_context(session, s["enterprise_id"])
    # 文件名兜底：主会话未给可读文件名（或给的是 slice_id 占位名）时，
    # 用该条目的模板标题（清理后）作默认名，避免产物退化成 slice_id.docx
    if not str(filename or "").strip() or filename == f"{slice_id}.docx":
        filename = _safe_filename(_clean_item_name(s.get("title") or "")) + ".docx"
    if not filename.endswith(".docx"):
        filename += ".docx"
    art = AgentArtifact(
        enterprise_id=s["enterprise_id"],
        project_id=s["project_id"],
        task_id=int(task_id),
        kind="item_docx",
        name=f"{dir_name}/{filename}",
        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        content=data,
    )
    session.add(art)
    await session.commit()
    await _set_rls_context(session, s["enterprise_id"])
    _SLICES.pop(slice_id, None)  # 封存后切片即失效（不可再改）
    # 信息信号（不拦截）：主会话/验收子 agent 自查用
    return {
        "artifact_id": art.id,
        "name": art.name,
        "bytes": len(data),
        "req_title": s.get("title") or "",
        "matched_title": s.get("matched_title") or "",
        "was_verified": bool(s.get("verified")),
    }


async def quote_xlsx(
    session: AsyncSession,
    enterprise_id: int,
    project_id: int,
    task_id: int,
    sheets: list[dict] | None,
) -> dict:
    """报价单 xlsx：主会话给出 sheets=[{name,rows}]，服务端确定性生成。"""
    from app.models.agent import AgentArtifact
    from app.services import export_service
    from app.services.task_service import _set_rls_context  # noqa: PLC0415

    data = export_service.xlsx_bytes({"sheets": sheets or []})
    await _set_rls_context(session, enterprise_id)
    art = AgentArtifact(
        enterprise_id=enterprise_id,
        project_id=int(project_id),
        task_id=int(task_id),
        kind="xlsx",
        name="价格文件/报价单.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=data,
    )
    session.add(art)
    await session.commit()
    await _set_rls_context(session, enterprise_id)
    return {"artifact_id": art.id, "name": art.name, "bytes": len(data)}


async def upload_artifact_file(
    session: AsyncSession,
    enterprise_id: int,
    project_id: int,
    task_id: int,
    name: str,
    data: bytes,
) -> dict:
    """整文件交付通道：Hermes 直接写好的完整交付文件（docx/xlsx/pdf）落库为封存产物。

    与切片填空路径并列可选——Hermes 可自产整文件（python-docx/openpyxl/matplotlib 配图），
    服务端不再介入文档内容生成；打包/清单/审计与切片产物同一套机制。"""
    import io as _io
    import zipfile as _zip

    from app.models.agent import AgentArtifact
    from app.services.task_service import _set_rls_context  # noqa: PLC0415

    name = str(name or "").strip()
    if not name:
        raise ValueError("缺少文件路径名（如 技术文件/（二）专项响应文件.docx）")
    if len(data) > 60 * 1024 * 1024:
        raise ValueError("文件超过 60MB 上限")
    stem = name.rsplit("/", 1)[-1]
    ext = stem.rsplit(".", 1)[-1].lower() if "." in stem else ""
    if ext == "docx":
        try:
            with _zip.ZipFile(_io.BytesIO(data)) as zf:
                if "word/document.xml" not in zf.namelist():
                    raise ValueError
        except Exception as exc:  # noqa: BLE001
            raise ValueError("docx 结构无效（不是有效的 Word 文件）") from exc
        kind = "item_docx"
        mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif ext == "xlsx":
        kind = "xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif ext == "pdf":
        if not data.startswith(b"%PDF-"):
            raise ValueError("pdf 文件头无效")
        kind = "pdf"
        mime = "application/pdf"
    else:
        raise ValueError("仅支持 docx/xlsx/pdf；其他格式请转档后上传")

    await _set_rls_context(session, enterprise_id)
    art = AgentArtifact(
        enterprise_id=enterprise_id,
        project_id=int(project_id),
        task_id=int(task_id),
        kind=kind,
        name=name,
        mime=mime,
        content=data,
    )
    session.add(art)
    await session.commit()
    await _set_rls_context(session, enterprise_id)
    return {"artifact_id": art.id, "name": art.name, "bytes": len(data), "kind": kind}


async def replace_artifact_file(
    session: AsyncSession,
    enterprise_id: int,
    project_id: int,
    task_id: int,
    artifact_id: int,
    data: bytes,
) -> dict:
    """覆盖修改已封存产物（Hermes 修完文件直接换内容）：artifact_id 与包内路径名不变，
    内容整体替换——随时可改，改完重新打包即可。"""
    import io as _io
    import zipfile as _zip

    from sqlalchemy import select as _sa_select

    from app.models.agent import AgentArtifact
    from app.services.task_service import _set_rls_context  # noqa: PLC0415

    if len(data) > 60 * 1024 * 1024:
        raise ValueError("文件超过 60MB 上限")
    await _set_rls_context(session, enterprise_id)
    art = await session.scalar(
        _sa_select(AgentArtifact).where(
            AgentArtifact.id == int(artifact_id),
            AgentArtifact.enterprise_id == enterprise_id,
            AgentArtifact.project_id == int(project_id),
        )
    )
    if art is None:
        raise ValueError("产物不存在或不属于本项目")
    stem = (art.name or "").rsplit("/", 1)[-1]
    ext = stem.rsplit(".", 1)[-1].lower() if "." in stem else ""
    if ext == "docx":
        try:
            with _zip.ZipFile(_io.BytesIO(data)) as zf:
                if "word/document.xml" not in zf.namelist():
                    raise ValueError
        except Exception as exc:  # noqa: BLE001
            raise ValueError("docx 结构无效（不是有效的 Word 文件）") from exc
    elif ext == "xlsx":
        pass
    elif ext == "pdf":
        if not data.startswith(b"%PDF-"):
            raise ValueError("pdf 文件头无效")
    else:
        raise ValueError("仅支持 docx/xlsx/pdf 产物覆盖")
    art.content = data
    await session.commit()
    await _set_rls_context(session, enterprise_id)
    return {"artifact_id": art.id, "name": art.name, "bytes": len(data), "replaced": True}


async def render_qa_artifact(
    session: AsyncSession,
    enterprise_id: int,
    project_id: int,
    artifact_id: int,
    task_id: int,
) -> dict:
    """渲染质检（等价 Codex 的 Word 逐页渲染）：docx 产物 → LibreOffice headless 转 PDF →
    PyMuPDF 逐页渲染 PNG + 空白页/页数/字符量统计。PNG 供 Hermes 用 vision 抽查版面问题
    （表格跨页/图片方向/断页）。返回 {pdf_path, page_count, blank_pages, png_dir, png_paths, per_page}。"""
    import os as _os
    import subprocess as _sp
    import time as _time

    from sqlalchemy import select as _sa_select

    from app.models.agent import AgentArtifact
    from app.services.task_service import _set_rls_context  # noqa: PLC0415

    await _set_rls_context(session, enterprise_id)
    art = await session.scalar(
        _sa_select(AgentArtifact).where(
            AgentArtifact.id == int(artifact_id),
            AgentArtifact.enterprise_id == enterprise_id,
            AgentArtifact.project_id == int(project_id),
        )
    )
    if art is None:
        raise ValueError("产物不存在或不属于本项目")
    stem = (art.name or "").rsplit("/", 1)[-1]
    if not stem.lower().endswith(".docx"):
        raise ValueError("渲染质检仅支持 docx 产物")
    work = f"/tmp/bidvolt_qa_{int(task_id)}_{int(_time.time())}"
    _os.makedirs(work + "/pages", exist_ok=True)
    src = work + "/" + stem
    with open(src, "wb") as f:
        f.write(art.content)
    env = dict(_os.environ)
    env["HOME"] = "/tmp"
    try:
        proc = _sp.run(
            ["soffice", "--headless", "--convert-to", "pdf", "--outdir", work, src,
             "-env:UserInstallation=file:///tmp/lo_profile_render"],
            capture_output=True, timeout=180, env=env,
        )
    except _sp.TimeoutExpired as exc:
        raise ValueError("渲染超时（180s）：文件可能过大或损坏") from exc
    pdf_path = work + "/" + stem.rsplit(".", 1)[0] + ".pdf"
    if proc.returncode != 0 or not _os.path.exists(pdf_path):
        raise ValueError(f"LibreOffice 渲染失败：{proc.stderr.decode('utf-8', 'replace')[-300:]}")
    try:
        import fitz  # PyMuPDF
    except ImportError as exc:
        raise ValueError("服务器缺少 PyMuPDF（pymupdf），无法渲染质检") from exc
    doc = fitz.open(pdf_path)
    per_page: list[dict] = []
    blank_pages: list[int] = []
    png_paths: list[str] = []
    for i, page in enumerate(doc, 1):
        text = page.get_text().strip()
        # 低分辨率采样判空白页：整页近白且文字量为 0
        pix = page.get_pixmap(dpi=50)
        samples = pix.samples
        n = len(samples)
        mean = sum(samples) / max(n, 1)
        var = sum((s - mean) ** 2 for s in samples[::7]) / max(len(samples[::7]), 1)
        blank = (mean > 248 and var < 6 and not text)
        png = work + f"/pages/page_{i:03d}.png"
        page.get_pixmap(dpi=110).save(png)
        png_paths.append(png)
        per_page.append({"page": i, "chars": len(text), "blank": blank})
        if blank:
            blank_pages.append(i)
    doc.close()
    return {
        "artifact_id": int(artifact_id),
        "name": art.name,
        "page_count": len(per_page),
        "blank_pages": blank_pages,
        "pdf_path": pdf_path,
        "png_dir": work + "/pages",
        "png_paths": png_paths,
        "per_page": per_page,
        "note": "PNG 供 vision 抽查版面（表格跨页/图片方向/断页）；空白页须回修重渲。",
    }


async def package_zip(
    session: AsyncSession,
    enterprise_id: int,
    project_id: int,
    task_id: int,
    artifact_ids: list[int] | None,
    draft_file_id: int | None,
) -> dict:
    """打包：把主会话已封存的条目文件/报价单 zip 成响应文件包，
    自动附 会话记录/主会话记录.md 与 manifest.json。返回 zip 产物 id。"""
    import io as _io
    import json as _json
    import zipfile as _zip

    from sqlalchemy import select as sa_select

    from app.models.agent import AgentArtifact
    from app.models.file import FileObject
    from app.models.task import Task
    from app.services import agent_pipeline
    from app.services.task_service import _set_rls_context  # noqa: PLC0415

    await _set_rls_context(session, enterprise_id)
    arts = (
        await session.scalars(
            sa_select(AgentArtifact).where(
                AgentArtifact.id.in_([int(i) for i in (artifact_ids or [])]),
                AgentArtifact.enterprise_id == enterprise_id,
                AgentArtifact.project_id == int(project_id),
                AgentArtifact.task_id == int(task_id),
                AgentArtifact.kind.in_(["item_docx", "xlsx", "pdf"]),
            )
        )
    ).all()
    if not arts:
        raise ValueError("没有可打包的封存文件（请先 seal_template_item / build_quote_xlsx）")

    draft_name = "采购文件"
    if draft_file_id:
        fobj = await session.get(FileObject, int(draft_file_id))
        if fobj is not None and fobj.enterprise_id == enterprise_id:
            draft_name = fobj.original_name or "采购文件"

    # ===== 打包前全量核对（信息信号，不拦截——产品决定：合规性由主会话+验收/评审子 agent 保证）=====
    # 1) 清单全覆盖：is_file_item 每条是否有对应封存文件进包；
    # 2) 同部分不雷同：每个目录内任意两份 docx 原文是否一致；
    # 3) 身份绑定：每份 docx 正文开头是否含自身条目标题。
    # 发现项如实返回/写入 manifest，由主会话决定回修后重新打包。
    import re as _re

    from lxml import etree as _etree

    from app.services.export_service import _elem_text, _item_key, canonical_text

    def _norm(s: str) -> str:
        return "".join(str(s).split())

    def _stem(a) -> str:
        stem = a.name.rsplit("/", 1)[-1]
        return stem[:-5] if stem.endswith(".docx") else stem

    item_arts = [a for a in arts if a.kind == "item_docx"]

    missing_items: list[str] = []
    dup_pairs: list[str] = []
    identity_issues: list[str] = []
    try:
        outline = await get_template_outline(session, enterprise_id, project_id)
        required = [
            it["title"]
            for items in outline["items_by_role"].values()
            for it in items
            if it.get("is_file_item")
        ]
        covered = {_item_key(_stem(a)) for a in item_arts}
        missing_items = [t for t in required if _item_key(t) not in covered]
    except Exception:  # noqa: BLE001 清单读取失败不阻塞打包
        logger.warning("打包完整性核对失败", exc_info=True)
        missing_items = []

    # 硬门禁（R8 教训：首轮打包缺 4 份 is_file_item 条目，靠 8 轮续跑才补齐）：
    # 清单未全覆盖直接拒绝打包并列出缺失条目——服务端保证结构完整，
    # 不再依赖主会话自觉。清单读取失败（unknown 状态）不拦截。
    if missing_items:
        raise ValueError(
            "以下 is_file_item 条目尚未 seal 进包，请逐份 slice→fill→append→verify→seal 后重新打包："
            + "；".join(missing_items[:12])
        )

    texts: dict[int, str] = {}
    dir_texts: dict[str, list[tuple[str, str]]] = {}
    for a in item_arts:
        try:
            with _zip.ZipFile(_io.BytesIO(a.content)) as zf:
                root = _etree.fromstring(zf.read("word/document.xml"))
            text = canonical_text(root)
        except Exception as exc:  # noqa: BLE001 解析失败如实记录，不阻塞
            identity_issues.append(f"{a.name}：无法解析（{exc}）")
            continue
        texts[a.id] = text
        dir_texts.setdefault(a.name.rsplit("/", 1)[0], []).append((a.name, _norm(text)))

    for _d, entries in dir_texts.items():
        for i in range(len(entries)):
            for j in range(i + 1, len(entries)):
                if entries[i][1] == entries[j][1]:
                    dup_pairs.append(f"{entries[i][0]} == {entries[j][0]}")

    for a in item_arts:
        key = _item_key(_stem(a))
        if a.id in texts and key and key not in _norm(texts[a.id])[:1200]:
            identity_issues.append(f"{a.name}：正文开头不含自身条目标题「{key}」")
    # 裸待补充信号：label 为空或「具体标签」字样的逐处列出。
    # 文本口径=最终文本：只收 w:t（插入层 w:ins 的 w:t 天然包含；删除层 w:delText 不是 w:t，天然排除）。
    # 曾用 root.itertext() 把删除层旧标记一并计入（fill 多轮后旧【待补充】落在 w:delText）→ 误报
    # 与 verify 的 remaining_blanks 口径不一致，逼得主会话自行解包核对（任务 380 教训）。
    bare_pending: dict[str, list] = {}
    for a in item_arts:
        try:
            with _zip.ZipFile(_io.BytesIO(a.content)) as zf:
                root = _etree.fromstring(zf.read("word/document.xml"))
            full = _elem_text(root)
        except Exception:  # noqa: BLE001
            continue
        items = []
        for m in _re.finditer(r"【待补充[^】]*】", full):
            label = m.group(0)[5:-1] if m.group(0).startswith("【待补充：") else ""
            if not label or "具体标签" in label:
                items.append({"label": label, "context": full[max(0, m.start() - 12):m.start()]})
        if items:
            bare_pending[a.name] = items
    # 表格/字体质量硬门禁（R10 教训：报价明细表用 | 竖线文本模拟表格、
    # 响应函中文 run 缺 eastAsia 字体导致 Windows 渲染错乱——服务端直接拒绝，
    # 不依赖主会话自觉；只对 docx 类条目检查，解析失败跳过）
    from app.services.export_service import _W_NS as _W_NS_Q  # noqa: PLC0415

    _W_Q = f"{{{_W_NS_Q}}}"
    docx_quality: dict[str, dict] = {}
    _pipe_row_re = _re.compile(r"^\s*\|.*\|\s*$")
    _cjk_re = _re.compile(r"[\u4e00-\u9fff]")
    # 只拦客观错误：中文 run 完全没设 eastAsia（Windows 按西文默认字体渲染中文=错乱），
    # 或把纯西文字体设成 eastAsia。宋体/仿宋/黑体/楷体/等线等中文字体一律放行——
    # 不强制特定字体，按模板原字体即可。
    _latin_fonts = {"Calibri", "Calibri Light", "Consolas", "Times New Roman", "Arial"}
    for a in item_arts:
        try:
            with _zip.ZipFile(_io.BytesIO(a.content)) as zf:
                root = _etree.fromstring(zf.read("word/document.xml"))
        except Exception:  # noqa: BLE001 非 docx 或解析失败跳过
            continue
        pipes = 0
        for p in root.iter(f"{_W_Q}p"):
            text = "".join(t.text or "" for t in p.iter(f"{_W_Q}t"))
            if text.strip() and _pipe_row_re.match(text):
                pipes += 1
        font_issues = 0
        for r in root.iter(f"{_W_Q}r"):
            text = "".join(t.text or "" for t in r.iter(f"{_W_Q}t"))
            if not text or not _cjk_re.search(text):
                continue
            rpr = r.find(f"{_W_Q}rPr")
            if rpr is None:
                font_issues += 1
                continue
            rf = rpr.find(f"{_W_Q}rFonts")
            east = rf.get(f"{_W_Q}eastAsia") if rf is not None else None
            if not east or east in _latin_fonts:
                font_issues += 1
        docx_quality[a.name] = {
            "tables": len(root.findall(f".//{_W_Q}tbl")),
            "pipe_paragraphs": pipes,
            "font_issues": font_issues,
        }
    pipe_files = [f"{n}（{q['pipe_paragraphs']} 段竖线假表）" for n, q in docx_quality.items() if q["pipe_paragraphs"]]
    font_files = [f"{n}（{q['font_issues']} 处字体不合规）" for n, q in docx_quality.items() if q["font_issues"]]
    if pipe_files:
        raise ValueError(
            "交付件存在竖线假表格（用 | 字符在正文模拟表格），请改为真实 Word 表格（w:tbl）后重新打包："
            + "；".join(pipe_files[:8])
        )
    if font_files:
        raise ValueError(
            "交付件字体不合规（中文 run 缺少中文字体设置、或把 Calibri/Consolas 等西文字体"
            "设成了中文字体——Windows 上会渲染错乱）。请为中文 run 显式设置 eastAsia 中文字体"
            "（按模板原字体，如宋体/仿宋/黑体）后重新打包："
            + "；".join(font_files[:8])
        )
    audit = {
        "checked": len(item_arts),
        "coverage_ok": not missing_items,
        "unique_ok": not dup_pairs,
        "identity_ok": not identity_issues,
        "missing_file_items": missing_items,
        "duplicate_pairs": dup_pairs,
        "identity_issues": identity_issues,
        "bare_pending": bare_pending,
        "bare_pending_count": sum(len(v) for v in bare_pending.values()),
        "docx_quality": docx_quality,
    }

    buf = _io.BytesIO()
    files_manifest = []
    seen: set[str] = set()
    with _zip.ZipFile(buf, "w", _zip.ZIP_DEFLATED) as zf:
        for a in arts:
            # 同名产物去重：zip 内同名条目会被解压器静默覆盖（内容丢失类缺陷），
            # 必须重命名为 名称(2).ext，绝不静默覆盖
            name = a.name
            k = 2
            while name in seen:
                if "." in name.rsplit("/", 1)[-1]:
                    stem_n, dot, ext = name.rpartition(".")
                    name = f"{stem_n}({k}).{ext}"
                else:
                    name = f"{name}({k})"
                k += 1
            seen.add(name)
            zf.writestr(name, a.content)
            files_manifest.append({"name": name, "bytes": len(a.content)})
        # 主会话全程记录（纯代码生成：完整版 + 精简版都由服务端从事件库渲染，主会话不感知）
        task = await session.scalar(sa_select(Task).where(Task.id == int(task_id)))
        if task is not None:
            try:
                record = await agent_pipeline.session_record_markdown(session, task)
                zf.writestr("会话记录/主会话记录.md", record.encode("utf-8"))
                files_manifest.append({"name": "会话记录/主会话记录.md", "bytes": len(record.encode("utf-8"))})
                condensed = agent_pipeline.condense_session_markdown(record)
                zf.writestr("会话记录/主会话记录-精简版.md", condensed.encode("utf-8"))
                files_manifest.append({"name": "会话记录/主会话记录-精简版.md", "bytes": len(condensed.encode("utf-8"))})
            except Exception:  # noqa: BLE001 会话记录缺失不影响打包主体
                logger.warning("打包附会话记录失败", exc_info=True)

        manifest = {
            "project_id": int(project_id),
            "task_id": int(task_id),
            "draft": draft_name,
            "note": "主会话经成文工具链（切片→填空→追加→校验→封存→打包）自主成文；"
                    "全部改动可在 Word【审阅→所有标记】中逐处查看；附主会话全程记录。",
            "audit": audit,
            "missing_file_items": missing_items,
            "files": files_manifest,
        }
        zf.writestr("manifest.json", _json.dumps(manifest, ensure_ascii=False, indent=2))
    data = buf.getvalue()

    art = AgentArtifact(
        enterprise_id=enterprise_id,
        project_id=int(project_id),
        task_id=int(task_id),
        kind="zip",
        name=f"响应文件包(底稿：{draft_name}).zip",
        mime="application/zip",
        content=data,
    )
    session.add(art)
    await session.commit()
    await _set_rls_context(session, enterprise_id)
    # 信息信号（不拦截）：主会话据此决定是否回修重打包
    return {
        "artifact_id": art.id,
        "name": art.name,
        "bytes": len(data),
        "file_count": len(files_manifest),
        "missing_file_items": missing_items,
        "audit": audit,
    }


async def list_artifacts(
    session: AsyncSession,
    enterprise_id: int,
    project_id: int,
    task_id: int,
) -> dict:
    """产物清单（产物自检）：本任务已封存的全部产物（条目 docx/报价单 xlsx/响应包 zip）。"""
    from sqlalchemy import select as sa_select

    from app.models.agent import AgentArtifact
    from app.services.task_service import _set_rls_context  # noqa: PLC0415

    await _set_rls_context(session, enterprise_id)
    rows = (
        await session.scalars(
            sa_select(AgentArtifact)
            .where(
                AgentArtifact.enterprise_id == enterprise_id,
                AgentArtifact.project_id == int(project_id),
                AgentArtifact.task_id == int(task_id),
            )
            .order_by(AgentArtifact.id)
        )
    ).all()
    return {
        "artifacts": [
            {
                "artifact_id": a.id,
                "kind": a.kind,
                "name": a.name,
                "bytes": len(a.content or b""),
                "created_at": a.created_at.isoformat() if a.created_at else None,
            }
            for a in rows
        ]
    }


async def inspect_artifact(
    session: AsyncSession,
    enterprise_id: int,
    project_id: int,
    task_id: int,
    artifact_id: int,
) -> dict:
    """产物自检：预览已封存产物的内容（docx 文本/修订残留计数/待补充计数；xlsx 表格预览；zip 文件清单），
    供主会话/验收子 agent 核对"导出产物"与"成果模型"是否一致——补上模型与产物之间的验证盲区。"""
    import io as _io
    import zipfile as _zip

    from sqlalchemy import select as sa_select

    from app.models.agent import AgentArtifact
    from app.services.task_service import _set_rls_context  # noqa: PLC0415

    await _set_rls_context(session, enterprise_id)
    art = await session.scalar(
        sa_select(AgentArtifact).where(
            AgentArtifact.id == int(artifact_id),
            AgentArtifact.enterprise_id == enterprise_id,
            AgentArtifact.project_id == int(project_id),
            AgentArtifact.task_id == int(task_id),
        )
    )
    if art is None:
        raise ValueError("产物不存在或不属于本任务")
    base = {
        "artifact_id": art.id,
        "kind": art.kind,
        "name": art.name,
        "bytes": len(art.content or b""),
    }
    if art.kind == "item_docx":
        import re as _re

        from lxml import etree as _etree

        from app.services.export_service import _W_NS, _elem_text, tables_inventory

        W = f"{{{_W_NS}}}"
        with _zip.ZipFile(_io.BytesIO(art.content)) as zf:
            doc = _etree.fromstring(zf.read("word/document.xml"))
        # 最终文本口径：只收 w:t（含插入层、不含删除层 w:delText——多轮 fill 后旧标记落在删除层，
        # itertext 会把它们算进来造成"audit 报裸、verify 报干净"的口径打架，任务 380 曾因此空转）
        text = _elem_text(doc)
        # 待补充逐项清单（信息信号）：让检查者一眼看到"哪些还没填、分别要补什么"，
        # 而不是只给一个计数——计数会掩盖"本可填实却空着/标签含混"的问题；
        # 裸待补充（label 空）与「具体标签」模板字样打 kind=bare 并排最前——验收判据最该先看它们
        pending_items: list[dict] = []
        for m in _re.finditer(r"【待补充[^】]*】", text):
            label = m.group(0)[5:-1] if m.group(0).startswith("【待补充：") else ""
            start = max(0, m.start() - 18)
            pending_items.append(
                {
                    "label": label,
                    "context": text[start:m.start()],
                    "kind": "bare" if (not label or "具体标签" in label) else "labeled",
                }
            )
        pending_items.sort(key=lambda x: 0 if x["kind"] == "bare" else 1)
        bare_count = sum(1 for x in pending_items if x["kind"] == "bare")
        base.update(
            {
                "text_preview_head": text[:600],
                "text_preview_tail": text[-300:],
                "chars": len(text),
                "pending_count": text.count("【待补充"),
                "pending_items": pending_items,
                "bare_pending_count": bare_count,
                "tables": tables_inventory(doc),
                "ins_count": len(doc.findall(".//" + W + "ins")),
                "del_count": len(doc.findall(".//" + W + "del")),
            }
        )
    elif art.kind == "xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(_io.BytesIO(art.content))
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            preview = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 4:
                    preview.append("…")
                    break
                preview.append([" " if c is None else str(c)[:40] for c in row])
            sheets.append({"name": name, "rows": ws.max_row, "cols": ws.max_column, "preview": preview})
        base["sheets"] = sheets
    elif art.kind == "zip":
        with _zip.ZipFile(_io.BytesIO(art.content)) as zf:
            base["entries"] = [{"name": n, "bytes": zf.getinfo(n).file_size} for n in zf.namelist()]
    else:
        base["note"] = "该产物类型无结构化预览"
    return base
